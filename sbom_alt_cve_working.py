#!/usr/bin/env python3
"""
sbom_alt_cve_working.py — ALT Linux CVE scanner for CycloneDX SBOM

Usage:
  python3 sbom_alt_cve_working.py inventory.cdx.json --c10f2 -o report.xlsx
  python3 sbom_alt_cve_working.py inventory.cdx.json --p10 --json
  python3 sbom_alt_cve_working.py --update-cache --c10f2

Requires:
  pip install requests openpyxl

Notes:
  - Scans ALT Linux RPM packages via ALT OVAL feed.
  - Matches OVAL entries by source package only to reduce false positives.
  - Exports a deliberately simple XLSX to avoid LibreOffice/Excel corruption.
  - Uses python rpm bindings if available; otherwise uses a pure Python rpmvercmp fallback.
"""

from __future__ import annotations

import argparse
import functools
import gzip
import io
import json
import re
import sys
import time
import zipfile
from pathlib import Path
from typing import Optional, Any
from urllib.parse import quote, unquote
import xml.etree.ElementTree as ET

try:
    import requests
except ImportError:
    print("Missing dependency: pip install requests openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except ImportError:
    print("Missing dependency: pip install requests openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import rpm as rpm_module  # type: ignore
except Exception:
    rpm_module = None

OVAL_API = "https://rdb.altlinux.org/api/errata/export/oval"
OVAL_CACHE_DIR = Path.home() / ".cache" / "alt_oval"
OVAL_CACHE_TTL = 86400
LATEST_CACHE_FILE = OVAL_CACHE_DIR / "latest_versions_cache.json"

DEF = "http://oval.mitre.org/XMLSchema/oval-definitions-5"
LNX = "http://oval.mitre.org/XMLSchema/oval-definitions-5#linux"

ALL_BRANCHES = ["p9", "p10", "p11", "c9f2", "c10f1", "c10f2"]
SUPPORTED_SCAN_BRANCHES = ["p9", "p10", "p11", "c9f2", "c10f2"]
UNSUPPORTED_BRANCHES = {"c10f1"}
TARGET_BRANCH: Optional[str] = None

PLATFORM_BRANCHES = {
    "p9": ["p9", "c9f2"],
    "p10": ["p10", "c10f2"],
    "p11": ["p11"],
    "p8": ["p8"],
}

_log_file = None


def init_log(path: str = "scan_cve.log") -> None:
    global _log_file
    _log_file = open(path, "w", encoding="utf-8")
    _log_file.write(f"# scan log {time.strftime('%Y-%m-%d %H:%M:%S')}\n\n")
    print(f"Verbose log: {path}", flush=True)


def log(msg: str) -> None:
    if _log_file:
        _log_file.write(msg + "\n")
        _log_file.flush()


_BRANCH_RE = re.compile(r"(?:^|[.\-_/:@+])(c\d+f\d+|p\d+)(?:[.\-_/:@+]|$)", re.I)
_DISTRO_RE = re.compile(r"[?&]distro=altlinux-(\d+)\b", re.I)
_PURL_BR_RE = re.compile(r"[?&]branch=(c\d+f\d+|p\d+)\b", re.I)


def find_branch_in_text(s: str) -> str:
    if not s:
        return ""
    m = _BRANCH_RE.search(s)
    return m.group(1).lower() if m else ""


def resolve_branch(comp: dict[str, Any]) -> dict[str, str]:
    for field in ("src_rpm", "release", "version"):
        b = find_branch_in_text(comp.get(field, ""))
        if b:
            return {"branch": b, "confidence": "exact", "evidence": {"src_rpm": "sourceRpm"}.get(field, field)}

    purl = comp.get("purl", "")
    m = _PURL_BR_RE.search(purl)
    if m:
        return {"branch": m.group(1).lower(), "confidence": "exact", "evidence": "purl.branch"}

    loc = comp.get("location_path", "")
    if loc:
        b = find_branch_in_text(unquote(loc))
        if b:
            return {"branch": b, "confidence": "exact", "evidence": "location"}

    m = _DISTRO_RE.search(purl)
    if m:
        return {"branch": f"p{m.group(1)}", "confidence": "platform", "evidence": "purl.distro"}

    return {"branch": "", "confidence": "unknown", "evidence": ""}


def branches_to_scan(comp: dict[str, Any]) -> list[str]:
    if TARGET_BRANCH:
        return [TARGET_BRANCH]

    conf = comp.get("branch_conf", "unknown")
    branch = comp.get("branch", "")

    if conf == "exact" and branch:
        return [branch]
    if conf == "platform" and branch:
        return PLATFORM_BRANCHES.get(branch, [branch])
    return [b for b in ALL_BRANCHES if b not in UNSUPPORTED_BRANCHES]


def get_oval(branch: str, no_cache: bool = False) -> str:
    OVAL_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache = OVAL_CACHE_DIR / f"{branch}.xml"

    if no_cache and cache.exists():
        cache.unlink(missing_ok=True)

    if cache.exists():
        age = time.time() - cache.stat().st_mtime
        if age < OVAL_CACHE_TTL:
            text = cache.read_text(encoding="utf-8", errors="replace")
            if text.lstrip().startswith("<"):
                print(f"  OVAL from cache [{branch}] ({int(age // 3600)}h)", flush=True)
                return text
            cache.unlink(missing_ok=True)

    url = f"{OVAL_API}/{branch}?one_file=true"
    print(f"  Downloading OVAL [{branch}]: {url}", flush=True)
    try:
        r = requests.get(url, timeout=120)
        r.raise_for_status()
    except Exception as e:
        print(f"  WARN: OVAL download failed [{branch}]: {e}", flush=True)
        return ""

    text = decompress_response(r.content, branch)
    if not text:
        return ""

    try:
        ET.fromstring(text)
    except ET.ParseError as e:
        print(f"  WARN: invalid OVAL XML [{branch}]: {e}", flush=True)
        return ""

    cache.write_text(text, encoding="utf-8")
    print(f"  OVAL saved [{branch}] ({len(text) // 1024} KiB)", flush=True)
    return text


def decompress_response(raw: bytes, branch: str) -> str:
    if raw.lstrip().startswith(b"<?xml") or raw.lstrip().startswith(b"<oval"):
        return raw.decode("utf-8", errors="replace")

    if raw[:2] == b"PK":
        try:
            with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                names = [n for n in zf.namelist() if n.endswith(".xml")]
                if not names:
                    print(f"  WARN: ZIP has no XML [{branch}]", flush=True)
                    return ""
                return zf.read(names[0]).decode("utf-8", errors="replace")
        except Exception as e:
            print(f"  WARN: ZIP error [{branch}]: {e}", flush=True)
            return ""

    if raw[:2] == b"\x1f\x8b":
        try:
            return gzip.decompress(raw).decode("utf-8", errors="replace")
        except Exception as e:
            print(f"  WARN: GZIP error [{branch}]: {e}", flush=True)
            return ""

    print(f"  WARN: unknown OVAL format [{branch}]: {raw[:16].hex()}", flush=True)
    return ""


def branch_from_platform(platform: str) -> str:
    m = re.search(r"\b(c\d+f\d+|p\d+)\b", platform or "", re.I)
    return m.group(1).lower() if m else ""


def parse_oval(xml_text: str) -> list[dict[str, Any]]:
    if not xml_text:
        return []

    root = ET.fromstring(xml_text)

    objs: dict[str, str] = {}
    for o in root.iter(f"{{{LNX}}}rpminfo_object"):
        el = o.find(f"{{{LNX}}}name")
        if el is not None and el.text:
            objs[o.get("id", "")] = el.text

    states: dict[str, str] = {}
    for s in root.iter(f"{{{LNX}}}rpminfo_state"):
        el = s.find(f"{{{LNX}}}evr")
        if el is not None and el.text:
            states[s.get("id", "")] = el.text

    tests: dict[str, tuple[str, str]] = {}
    for t in root.iter(f"{{{LNX}}}rpminfo_test"):
        o_ref = t.find(f"{{{LNX}}}object")
        s_ref = t.find(f"{{{LNX}}}state")
        if o_ref is not None and s_ref is not None:
            tests[t.get("id", "")] = (o_ref.get("object_ref", ""), s_ref.get("state_ref", ""))

    defs = root.find(f"{{{DEF}}}definitions")
    if defs is None:
        return []

    entries: list[dict[str, Any]] = []

    for defn in defs.findall(f"{{{DEF}}}definition"):
        if defn.get("class") != "patch":
            continue

        meta = defn.find(f"{{{DEF}}}metadata")
        if meta is None:
            continue

        entry: dict[str, Any] = {
            "src_package": "",
            "fixed_version": "",
            "severity": "",
            "branches": [],
            "bin_packages": [],
            "vulns": [],
            "title": "",
        }

        title_el = meta.find(f"{{{DEF}}}title")
        title = title_el.text if title_el is not None and title_el.text else ""
        entry["title"] = title
        if title:
            m = re.search(r"package\s+`([^`]+)`", title)
            if m:
                entry["src_package"] = m.group(1)
            m = re.search(r"to version\s+(\S+)", title)
            if m:
                entry["fixed_version"] = m.group(1)

        affected = meta.find(f"{{{DEF}}}affected")
        if affected is not None:
            platforms = [p.text for p in affected.findall(f"{{{DEF}}}platform") if p.text]
            entry["branches"] = sorted({branch_from_platform(p) for p in platforms if branch_from_platform(p)})

        adv = meta.find(f"{{{DEF}}}advisory")
        if adv is not None:
            sev = adv.find(f"{{{DEF}}}severity")
            if sev is not None:
                entry["severity"] = sev.text or ""
            for tag, source in (("cve", "CVE"), ("bdu", "BDU")):
                for el in adv.findall(f"{{{DEF}}}{tag}"):
                    vid = (el.text or "").strip()
                    if vid:
                        entry["vulns"].append({
                            "id": vid,
                            "source": source,
                            "cvss3": el.get("cvss3", ""),
                            "cvss2": el.get("cvss", ""),
                            "impact": el.get("impact", ""),
                            "href": el.get("href", ""),
                            "public": el.get("public", ""),
                        })

        criteria = defn.find(f"{{{DEF}}}criteria")
        if criteria is not None:
            bins: set[str] = set()
            fixed_evr = None
            for crit in criteria.iter(f"{{{DEF}}}criterion"):
                tref = crit.get("test_ref", "")
                if tref in tests:
                    oid, sid = tests[tref]
                    if objs.get(oid):
                        bins.add(objs[oid])
                    if fixed_evr is None and states.get(sid):
                        fixed_evr = states[sid]
            entry["bin_packages"] = sorted(bins)
            if not entry["fixed_version"] and fixed_evr:
                entry["fixed_version"] = re.sub(r"^\d+:", "", fixed_evr)

        if entry["src_package"] and entry["vulns"]:
            entries.append(entry)

    return entries


def prop(c: dict[str, Any], name: str) -> str:
    for p in c.get("properties", []) or []:
        if p.get("name") == name:
            return p.get("value", "") or ""
    return ""


def src_name_from_rpm(src_rpm: str) -> str:
    name = re.sub(r"\.src\.rpm$", "", src_rpm or "", flags=re.I)
    name = re.sub(r"-\d[\w.+~^]*(?:-[\w.+~^]+)*$", "", name)
    return name


def guess_ecosystem(purl: str, src_rpm: str, buildhost: str) -> str:
    p = (purl or "").lower()
    sr = (src_rpm or "").lower()
    bh = (buildhost or "").lower()
    if "pkg:rpm/" in p:
        if "altlinux" in bh or ".alt" in sr:
            return "ALT RPM"
        return "RPM"
    for prefix, name in (
        ("pkg:npm/", "npm"), ("pkg:pypi/", "PyPI"), ("pkg:maven/", "Maven"),
        ("pkg:golang/", "Go"), ("pkg:cargo/", "Rust/Cargo"), ("pkg:gem/", "RubyGems"),
        ("pkg:nuget/", "NuGet"), ("pkg:deb/", "Debian/Ubuntu"),
    ):
        if prefix in p:
            return name
    return "Unknown"


def parse_component(c: dict[str, Any]) -> Optional[dict[str, Any]]:
    name = c.get("name", "") or ""
    if not name or c.get("type") == "operating-system":
        return None

    purl = c.get("purl", "") or ""
    pkg_type = prop(c, "syft:package:type") or c.get("type", "") or ""
    version = re.sub(r"^\d+:", "", c.get("version", "") or "")
    src_rpm = prop(c, "syft:metadata:sourceRpm")
    release = prop(c, "syft:metadata:release")
    buildhost = prop(c, "rpm:buildhost")
    location_path = prop(c, "syft:location:0:path")

    if not src_rpm and "upstream=" in purl:
        m = re.search(r"upstream=([^&]+)", purl)
        if m:
            src_rpm = unquote(m.group(1))

    src_name = src_name_from_rpm(src_rpm) if src_rpm else ""
    is_rpm = purl.startswith("pkg:rpm/") or pkg_type == "rpm"
    gost_provided = prop(c, "GOST:provided_by").strip().lower() == "alt linux"

    if not is_rpm:
        return {
            "name": name, "version": version, "src_name": src_name, "src_rpm": src_rpm,
            "release": release, "purl": purl, "buildhost": buildhost, "is_alt": False,
            "no_buildhost": False, "ecosystem": guess_ecosystem(purl, src_rpm, buildhost),
            "branch": "", "branch_conf": "unknown", "branch_src": "", "gost_provided": False,
        }

    is_alt = bool(re.search(r"\.altlinux\.org$", buildhost, re.I)) or gost_provided
    no_buildhost = not buildhost and not gost_provided

    br = resolve_branch({
        "src_rpm": src_rpm,
        "release": release,
        "version": version,
        "purl": purl,
        "location_path": location_path,
    })

    return {
        "name": name,
        "version": version,
        "src_name": src_name,
        "src_rpm": src_rpm,
        "release": release,
        "purl": purl,
        "buildhost": buildhost,
        "is_alt": is_alt,
        "no_buildhost": no_buildhost,
        "ecosystem": guess_ecosystem(purl, src_rpm, buildhost),
        "branch": br["branch"],
        "branch_conf": br["confidence"],
        "branch_src": br["evidence"],
        "gost_provided": gost_provided,
    }


def read_sbom(path: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    alt: list[dict[str, Any]] = []
    other: list[dict[str, Any]] = []
    no_bh: list[dict[str, Any]] = []

    def walk(comps: list[dict[str, Any]]) -> None:
        for c in comps or []:
            comp = parse_component(c)
            if comp:
                if comp["is_alt"]:
                    alt.append(comp)
                elif comp["no_buildhost"]:
                    no_bh.append(comp)
                else:
                    other.append(comp)
            if "components" in c:
                walk(c.get("components") or [])

    walk(data.get("components", []) or [])
    return alt, other, no_bh


def split_evr(evr: str) -> tuple[str, str, str]:
    evr = evr or ""
    epoch = "0"
    if ":" in evr:
        epoch, evr = evr.split(":", 1)
        epoch = epoch or "0"
    version = evr
    release = ""
    if "-" in evr:
        version, release = evr.rsplit("-", 1)
    return str(epoch), version, release


def rpmvercmp_segment(a: str, b: str) -> int:
    """Pure Python approximation of rpmvercmp, handles alnum, ~ and ^."""
    if a == b:
        return 0
    i = j = 0
    la, lb = len(a), len(b)

    while i < la or j < lb:
        while i < la and not a[i].isalnum() and a[i] not in "~^":
            i += 1
        while j < lb and not b[j].isalnum() and b[j] not in "~^":
            j += 1

        if i < la and a[i] == "~":
            if j >= lb or b[j] != "~":
                return -1
            i += 1; j += 1
            continue
        if j < lb and b[j] == "~":
            return 1

        if i < la and a[i] == "^":
            if j >= lb:
                return 1
            if b[j] != "^":
                return 1
            i += 1; j += 1
            continue
        if j < lb and b[j] == "^":
            if i >= la:
                return -1
            return -1

        if i >= la or j >= lb:
            break

        isnum_a = a[i].isdigit()
        isnum_b = b[j].isdigit()

        ia = i
        if isnum_a:
            while i < la and a[i].isdigit():
                i += 1
            seg_a = a[ia:i].lstrip("0") or "0"
        else:
            while i < la and a[i].isalpha():
                i += 1
            seg_a = a[ia:i]

        jb = j
        if isnum_b:
            while j < lb and b[j].isdigit():
                j += 1
            seg_b = b[jb:j].lstrip("0") or "0"
        else:
            while j < lb and b[j].isalpha():
                j += 1
            seg_b = b[jb:j]

        if isnum_a and not isnum_b:
            return 1
        if not isnum_a and isnum_b:
            return -1

        if isnum_a and len(seg_a) != len(seg_b):
            return 1 if len(seg_a) > len(seg_b) else -1

        if seg_a != seg_b:
            return 1 if seg_a > seg_b else -1

    if i >= la and j >= lb:
        return 0
    if i >= la:
        return -1
    return 1


def rpm_cmp(a: str, b: str) -> int:
    ea, va, ra = split_evr(a)
    eb, vb, rb = split_evr(b)

    if rpm_module is not None and hasattr(rpm_module, "labelCompare"):
        try:
            return int(rpm_module.labelCompare((ea, va, ra), (eb, vb, rb)))
        except Exception:
            pass

    try:
        ea_i, eb_i = int(ea), int(eb)
        if ea_i != eb_i:
            return 1 if ea_i > eb_i else -1
    except ValueError:
        c = rpmvercmp_segment(ea, eb)
        if c:
            return c

    c = rpmvercmp_segment(va, vb)
    if c:
        return c
    return rpmvercmp_segment(ra, rb)


def ver_lt(a: str, b: str) -> bool:
    return rpm_cmp(a, b) < 0


def rpm_max(values: list[str]) -> str:
    if not values:
        return ""
    return max(values, key=functools.cmp_to_key(rpm_cmp))


def selftest_rpm_compare() -> None:
    tests = [
        ("1.1.1w-alt0.p10.2", "1.1.1d-alt1", 1),
        ("2.9.12-alt1.p10.1", "2.9.12-alt1.p10.8", -1),
        ("7.9p1-alt4.p10.7", "7.9p1-alt4.p10.8", -1),
    ]
    for a, b, expected in tests:
        got = rpm_cmp(a, b)
        if expected > 0 and got <= 0:
            raise RuntimeError(f"RPM compare broken: {a} should be newer than {b}; got {got}")
        if expected < 0 and got >= 0:
            raise RuntimeError(f"RPM compare broken: {a} should be older than {b}; got {got}")


def scan(components: list[dict[str, Any]], oval_entries: list[dict[str, Any]], debug: bool = False) -> list[dict[str, Any]]:
    by_src: dict[str, list[dict[str, Any]]] = {}
    by_name: dict[str, list[dict[str, Any]]] = {}

    for comp in components:
        src = comp.get("src_name") or ""
        if src:
            by_src.setdefault(src, []).append(comp)
        by_name.setdefault(comp["name"], []).append(comp)

    result: dict[str, dict[str, Any]] = {}

    for entry in oval_entries:
        src_pkg = entry.get("src_package", "")
        if not src_pkg:
            continue

        e_branches = set(entry.get("branches", []) or [])
        candidates = by_src.get(src_pkg, []) or by_name.get(src_pkg, [])

        for comp in candidates:
            scan_branches = branches_to_scan(comp)
            target = [b for b in scan_branches if not e_branches or b in e_branches]
            if not target:
                continue

            fix = entry.get("fixed_version", "") or ""
            if fix and not ver_lt(comp["version"], fix):
                if debug and src_pkg in {"openssl1.1", "openssl", "libressl"}:
                    log(f"SKIP {comp['name']} installed={comp['version']} fix={fix} cmp={rpm_cmp(comp['version'], fix)}")
                continue

            comp_name = comp["name"]
            if comp_name not in result:
                result[comp_name] = {"component": comp, "findings_by_branch": {}}

            for branch in target:
                for vuln in entry.get("vulns", []):
                    result[comp_name]["findings_by_branch"].setdefault(branch, []).append({
                        "vuln": vuln,
                        "fixed_version": fix,
                        "severity": entry.get("severity", ""),
                        "branch": branch,
                        "oval_src_package": src_pkg,
                        "oval_title": entry.get("title", ""),
                    })

    output: list[dict[str, Any]] = []
    for data in result.values():
        flat: list[dict[str, Any]] = []
        clean_by_branch: dict[str, str] = {}
        for branch, findings in data["findings_by_branch"].items():
            seen: dict[str, dict[str, Any]] = {}
            for f in findings:
                vid = f["vuln"]["id"]
                if vid not in seen or rpm_cmp(f["fixed_version"], seen[vid]["fixed_version"]) > 0:
                    seen[vid] = f
            deduped = list(seen.values())
            fixes = [f["fixed_version"] for f in deduped if f["fixed_version"]]
            clean_by_branch[branch] = rpm_max(fixes) if fixes else ""
            flat.extend(deduped)

        if flat:
            data["findings"] = flat
            data["clean_ver_by_branch"] = clean_by_branch
            output.append(data)

    return output


def load_latest_cache() -> dict[str, Any]:
    try:
        if LATEST_CACHE_FILE.exists():
            return json.loads(LATEST_CACHE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return {}


def save_latest_cache(cache: dict[str, Any]) -> None:
    try:
        OVAL_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        tmp = LATEST_CACHE_FILE.with_suffix(".tmp")
        tmp.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(LATEST_CACHE_FILE)
    except Exception:
        pass


def fetch_package_versions(src_name: str, debug: bool = False) -> dict[str, str]:
    url = f"https://rdb.altlinux.org/api/site/package_versions?name={quote(src_name)}"
    last_error = ""
    for attempt in range(1, 4):
        try:
            r = requests.get(url, timeout=30)
            if r.status_code == 200:
                bmap: dict[str, str] = {}
                for v in r.json().get("versions", []) or []:
                    b = (v.get("branch") or "").lower()
                    ver = v.get("version") or ""
                    rel = v.get("release") or ""
                    if b and ver:
                        bmap[b] = f"{ver}-{rel}" if rel else ver
                return bmap
            last_error = f"HTTP {r.status_code}"
        except Exception as e:
            last_error = str(e)
        time.sleep(0.5 * attempt)
    if debug:
        log(f"package_versions {src_name}: failed after retries: {last_error}")
    return {}


def fetch_latest_versions(results: list[dict[str, Any]], debug: bool = False) -> dict[str, dict[str, str]]:
    cache = load_latest_cache()
    result: dict[str, dict[str, str]] = {}
    total = len(results)
    found = 0
    misses = 0
    print("\nFetching latest versions...", flush=True)

    actual_branches = {TARGET_BRANCH} if TARGET_BRANCH else {"sisyphus", "p9", "p10", "p11", "c9f2", "c10f2"}

    for data in results:
        comp = data["component"]
        src_name = comp.get("src_name") or comp["name"]
        comp_name = comp["name"]

        if src_name not in cache or not cache.get(src_name):
            bmap = fetch_package_versions(src_name, debug=debug)
            if bmap:
                cache[src_name] = bmap
                save_latest_cache(cache)
            else:
                misses += 1
                cache.setdefault(src_name, {})

        per_branch = {b: v for b, v in (cache.get(src_name) or {}).items() if b in actual_branches and v}
        if per_branch:
            result[comp_name] = per_branch
            found += 1

    print(f"  Got latest versions: {found}/{total}", flush=True)
    if misses:
        print(f"  WARN: {misses} package version lookups failed", flush=True)
    return result


def xlsx_safe(value: Any) -> Any:
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    value = ILLEGAL_CHARACTERS_RE.sub("", value)
    if len(value) > 32000:
        value = value[:31997] + "..."
    return value


def join_branch_map(d: dict[str, str]) -> str:
    if not d:
        return ""
    return "\n".join(f"{k}: {v}" for k, v in sorted(d.items()))


def branch_label(comp: dict[str, Any]) -> str:
    if TARGET_BRANCH:
        real = comp.get("branch") or ""
        if real and real != TARGET_BRANCH and comp.get("branch_conf") == "exact":
            return f"{TARGET_BRANCH} forced; source says {real}"
        return f"{TARGET_BRANCH} forced"
    return comp.get("branch") or "unknown"


def export_xlsx_safe(
    results: list[dict[str, Any]],
    total_components: int,
    out_path: str,
    non_alt: list[dict[str, Any]],
    no_buildhost: list[dict[str, Any]],
    latest_versions: dict[str, dict[str, str]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["metric", "value"])
    ws.append(["components_total", total_components])
    ws.append(["vulnerable_components", len(results)])
    ws.append(["findings_total", sum(len(r.get("findings", [])) for r in results)])
    ws.append(["non_alt_components", len(non_alt)])
    ws.append(["rpm_without_buildhost", len(no_buildhost)])
    ws.append(["target_branch", TARGET_BRANCH or "auto"])
    ws.append([])

    headers = [
        "package", "version", "source_rpm",
        "buildhost", "gost_provider", "ecosystem",
        "latest_by_branch",
        "max_severity", "findings_cve", "vuln_ids",
    ]
    ws.append(headers)

    sev_rank = {"critical": 4, "high": 3, "medium": 2, "low": 1, "": 0}

    for data in sorted(results, key=lambda x: x["component"]["name"]):
        comp = data["component"]
        findings = data.get("findings", [])
        max_sev = max((f.get("severity", "") for f in findings), key=lambda sev: sev_rank.get((sev or "").lower(), 0), default="")
        vuln_ids = ", ".join(sorted({f["vuln"]["id"] for f in findings}))
        conflict = "yes" if (TARGET_BRANCH and comp.get("branch_conf") == "exact" and comp.get("branch") and comp.get("branch") != TARGET_BRANCH) else ""
        ws.append([
            xlsx_safe(comp.get("name", "")), xlsx_safe(comp.get("version", "")),
            xlsx_safe(comp.get("src_rpm", "")),
            xlsx_safe(comp.get("buildhost", "")),
            xlsx_safe("yes" if comp.get("gost_provided") else ""), xlsx_safe(comp.get("ecosystem", "")),
            xlsx_safe(join_branch_map(latest_versions.get(comp.get("name", ""), {}))),
            xlsx_safe((max_sev or "").upper()), len(findings), xlsx_safe(vuln_ids),
        ])

    ws2 = wb.create_sheet("Details")
    ws2.append([
        "package", "version", "source_package", "source_rpm", "release", "buildhost",
        "gost_provided_by", "ecosystem", "detected_branch", "branch_confidence",
        "branch_evidence", "scan_branch", "branch_conflict",
        "vuln_id", "vuln_source", "severity", "fixed_version", "cvss", "href",
        "oval_src_package", "oval_title",
    ])

    for data in sorted(results, key=lambda x: x["component"]["name"]):
        comp = data["component"]
        conflict = "yes" if (TARGET_BRANCH and comp.get("branch_conf") == "exact" and comp.get("branch") and comp.get("branch") != TARGET_BRANCH) else ""
        for f in sorted(data.get("findings", []), key=lambda y: (y.get("branch", ""), y["vuln"]["id"])):
            v = f["vuln"]
            ws2.append([
                xlsx_safe(comp.get("name", "")), xlsx_safe(comp.get("version", "")),
                xlsx_safe(comp.get("src_name") or comp.get("name", "")), xlsx_safe(comp.get("src_rpm", "")),
                xlsx_safe(comp.get("release", "")), xlsx_safe(comp.get("buildhost", "")),
                xlsx_safe("yes" if comp.get("gost_provided") else ""), xlsx_safe(comp.get("ecosystem", "")),
                xlsx_safe(comp.get("branch", "")), xlsx_safe(comp.get("branch_conf", "")), xlsx_safe(comp.get("branch_src", "")),
                xlsx_safe(f.get("branch", "")), xlsx_safe(conflict), xlsx_safe(v.get("id", "")),
                xlsx_safe(v.get("source", "")), xlsx_safe(f.get("severity", "")), xlsx_safe(f.get("fixed_version", "")),
                xlsx_safe(v.get("cvss3") or v.get("cvss2") or ""), xlsx_safe(v.get("href", "")),
                xlsx_safe(f.get("oval_src_package", "")), xlsx_safe(f.get("oval_title", "")),
            ])

    ws3 = wb.create_sheet("Non-ALT components")
    ws3.append(["package", "version", "ecosystem", "purl", "source_package", "source_rpm", "release", "buildhost", "reason"])
    for comp in sorted(non_alt, key=lambda c: (c.get("ecosystem", ""), c.get("name", ""))):
        ws3.append([
            xlsx_safe(comp.get("name", "")), xlsx_safe(comp.get("version", "")), xlsx_safe(comp.get("ecosystem", "")),
            xlsx_safe(comp.get("purl", "")), xlsx_safe(comp.get("src_name", "")), xlsx_safe(comp.get("src_rpm", "")),
            xlsx_safe(comp.get("release", "")), xlsx_safe(comp.get("buildhost", "")),
            xlsx_safe("not ALT RPM or buildhost is not *.altlinux.org"),
        ])

    ws4 = wb.create_sheet("RPM without buildhost")
    ws4.append([
        "package", "version", "ecosystem", "purl", "source_package", "source_rpm", "release", "buildhost",
        "detected_branch", "branch_confidence", "branch_evidence", "reason",
    ])
    for comp in sorted(no_buildhost, key=lambda c: c.get("name", "")):
        ws4.append([
            xlsx_safe(comp.get("name", "")), xlsx_safe(comp.get("version", "")), xlsx_safe(comp.get("ecosystem", "")),
            xlsx_safe(comp.get("purl", "")), xlsx_safe(comp.get("src_name", "")), xlsx_safe(comp.get("src_rpm", "")),
            xlsx_safe(comp.get("release", "")), xlsx_safe(comp.get("buildhost", "")), xlsx_safe(comp.get("branch", "")),
            xlsx_safe(comp.get("branch_conf", "")), xlsx_safe(comp.get("branch_src", "")),
            xlsx_safe("RPM has no rpm:buildhost and no GOST:provided_by=ALT Linux"),
        ])

    out = Path(out_path)
    tmp = out.with_name(out.name + ".tmp.xlsx")
    if tmp.exists():
        tmp.unlink()
    wb.save(tmp)
    try:
        with zipfile.ZipFile(tmp, "r") as zf:
            bad = zf.testzip()
            if bad:
                raise RuntimeError(f"bad file inside xlsx: {bad}")
    except Exception as e:
        tmp.unlink(missing_ok=True)
        raise RuntimeError(f"XLSX validation failed before final save: {e}")
    tmp.replace(out)
    print(f"Excel saved: {out}", flush=True)


def build_json_output(results: list[dict[str, Any]], latest: dict[str, dict[str, str]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for data in results:
        comp = data["component"]
        out.append({
            "package": comp["name"],
            "version": comp["version"],
            "src_package": comp.get("src_name", ""),
            "source_rpm": comp.get("src_rpm", ""),
            "release": comp.get("release", ""),
            "buildhost": comp.get("buildhost", ""),
            "ecosystem": comp.get("ecosystem", ""),
            "gost_provided": bool(comp.get("gost_provided")),
            "branch": TARGET_BRANCH or comp.get("branch", ""),
            "detected_branch": comp.get("branch", ""),
            "branch_forced": bool(TARGET_BRANCH),
            "branch_conf": comp.get("branch_conf", ""),
            "branch_src": comp.get("branch_src", ""),
            "clean_ver_by_branch": data.get("clean_ver_by_branch", {}),
            "latest_by_branch": latest.get(comp["name"], {}),
            "vulns": [
                {
                    "package_buildhost": comp.get("buildhost", ""),
                    "id": f["vuln"].get("id", ""),
                    "source": f["vuln"].get("source", ""),
                    "branch": f.get("branch", ""),
                    "severity": f.get("severity", ""),
                    "fixed_version": f.get("fixed_version", ""),
                    "cvss3": f["vuln"].get("cvss3", ""),
                    "href": f["vuln"].get("href", ""),
                    "oval_src_package": f.get("oval_src_package", ""),
                }
                for f in data.get("findings", [])
            ],
        })
    return out


def main() -> None:
    global TARGET_BRANCH

    parser = argparse.ArgumentParser(description="ALT Linux SBOM CVE scanner using ALT OVAL")
    parser.add_argument("sbom", nargs="?", help="CycloneDX JSON SBOM path")
    parser.add_argument("-o", default="report.xlsx", help="output XLSX path")
    parser.add_argument("--json", action="store_true", help="print JSON instead of XLSX")
    parser.add_argument("--verbose", action="store_true", help="write scan_cve.log")
    parser.add_argument("--no-cache", action="store_true", help="ignore cached OVAL")
    parser.add_argument("--update-cache", action="store_true", help="update OVAL cache and exit")

    bg = parser.add_mutually_exclusive_group()
    for b in SUPPORTED_SCAN_BRANCHES:
        bg.add_argument(f"--{b}", action="store_true", help=f"scan only branch {b}")

    args = parser.parse_args()
    TARGET_BRANCH = next((b for b in SUPPORTED_SCAN_BRANCHES if getattr(args, b, False)), None)

    selftest_rpm_compare()

    if args.verbose:
        init_log()

    if TARGET_BRANCH:
        print(f"Target branch forced: {TARGET_BRANCH}", flush=True)

    if args.update_cache:
        branches = [TARGET_BRANCH] if TARGET_BRANCH else ALL_BRANCHES
        print(f"Updating OVAL cache: {', '.join(branches)}", flush=True)
        for b in branches:
            get_oval(b, no_cache=True)
        print("Done.", flush=True)
        return

    if not args.sbom:
        parser.print_help()
        sys.exit(1)

    print(f"Reading SBOM: {args.sbom}", flush=True)
    alt, other, no_bh = read_sbom(args.sbom)
    print(f"ALT RPM components:      {len(alt)}", flush=True)
    print(f"Other components:        {len(other)}", flush=True)
    print(f"RPM without buildhost:   {len(no_bh)}", flush=True)
    if no_bh:
        print(f"WARN: {len(no_bh)} RPM components without buildhost are not scanned as ALT", flush=True)

    total = len(alt) + len(other) + len(no_bh)
    if not alt:
        print("No ALT RPM components found.", flush=True)
        if not args.json:
            export_xlsx_safe([], total, args.o, other, no_bh, {})
        return

    branches = [TARGET_BRANCH] if TARGET_BRANCH else ALL_BRANCHES
    print(f"\nLoading OVAL branches: {', '.join(branches)}", flush=True)
    entries: list[dict[str, Any]] = []
    seen: set[tuple[Any, ...]] = set()
    for b in branches:
        xml = get_oval(b, no_cache=args.no_cache)
        parsed = parse_oval(xml)
        print(f"  [{b}] OVAL entries: {len(parsed)}", flush=True)
        for e in parsed:
            key = (e.get("src_package"), e.get("fixed_version"), tuple(sorted(v.get("id", "") for v in e.get("vulns", []))), tuple(e.get("branches", [])))
            if key not in seen:
                seen.add(key)
                entries.append(e)

    print(f"Unique OVAL entries: {len(entries)}", flush=True)
    results = scan(alt, entries, debug=args.verbose)
    print(f"Vulnerable components: {len(results)}", flush=True)

    latest = fetch_latest_versions(results, debug=args.verbose) if results else {}

    if args.json:
        print(json.dumps(build_json_output(results, latest), ensure_ascii=False, indent=2), flush=True)
    else:
        export_xlsx_safe(results, total, args.o, other, no_bh, latest)
        print(f"Done -> {args.o}", flush=True)


if __name__ == "__main__":
    main()
